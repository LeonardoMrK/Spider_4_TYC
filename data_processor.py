#!/usr/bin/env python
# coding: utf-8
# @Author  : Mr.K
# @Software: PyCharm Community Edition
# @Time    : 2019/12/17 9:46
# @Description: #对抓取下来的关键词为“教育”，“教育科技”，“教育咨询”的企业信息继续处理，实现：
#1.在经营范围中进行数据筛选，关键词为“教育咨询”，“培训”
#2.按照公司省份地址进行分sheet保存
#3.根据统计信息生成图表

import re
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from collections import defaultdict
plt.rcParams['font.sans-serif'] = ['SimHei']#绘图时显示中文
#plt.rcParams['axes.unicode_minus'] = False


data_file_path='E:\pycharmworkspace\Spider_4_TYC-master\input\comerge_data.xlsx'#所有数据，已经去重
save_file_path='E:\pycharmworkspace\Spider_4_TYC-master\output\output_comerge_data.xlsx'


# sheets = book.sheetnames# 从工作薄中获取一个表单(sheet)对象
# print(sheets, type(sheets))
# mySheet = book.create_sheet('mySheet')# 创建一个表单
# print(book.sheetnames)#此时尚未写入文件

def read_data(data_file_path):
    """
     数据预处理：从excel中读取信息并处理
    :param data_file_path: 文件路径
    :return: 处理后的信息
    """
    result_list=[]#将从excel中读取的数据存为list
    book=openpyxl.load_workbook(data_file_path)# 打开excel文件,获取工作簿对象
    activted_sheet = book.active#获取档期活动表单
    #获取想要的列
    colA = activted_sheet['A'] # 公司名称
    colF = activted_sheet['F'] # 省
    colG = activted_sheet['G'] # 市(有可能为空)
    colK = activted_sheet['K'] # 电话（有可能为空）
    colP = activted_sheet['P'] # 经营范围
    # for each in colK:
    #     print(type(each.value))
    #信息处理
    for i, (name, province,city,tele,domain) in enumerate(zip(colA,colF,colG,colK,colP)):
        result_str=str(i)+'$'+name.value+'$'+province.value+'$'+city.value+'$'+tele.value+'$'+domain.value
        result_list.append(result_str)
    print('raw_data总数:',len(result_list))
    return result_list


def filter_by_tele(result_list):
    """
    #电话号码筛选：把没有手机号码或者是座机号码的信息筛出去
    :param result_list: 从excel中读取并处理后的信息
    :return:筛选后的数据
    """
    final_result = []  # 最终筛选后的数据
    for i in range(len(result_list)):#依次获取每组数据
        temp_list=result_list[i].split('$')#将每一组数据split并重新存为list格式,每行5个属性，分别为0index，1name，2province，3city，4tele，5domain
        if temp_list[4]!='None' and len(temp_list[4])==11:#保存非空号码以及手机号码对应的数据
            final_result.append(result_list[i])
    print('after_filted_by_tele总数',len(final_result))  # 筛选后一共2609组数据
    return final_result



def filter_by_domain(result_list):
    """
    经营范围筛选：从domain属性中用正则匹配看是否含有关键词：“教育咨询”或“培训”
    :param result_list: 从excel中读取并处理后的信息
    :return:筛选后的数据
    """
    final_result = []  # 最终筛选后的数据
    for i in range(len(result_list)):  # 依次获取每组数据
        temp_list = result_list[i].split(
            '$')  # 将每一组数据split并重新存为list格式,每行5个属性，分别为0index，1name，2province，3city，4tele，5domain
        flag1 = re.findall('教育咨询',temp_list[5])#从domain属性中用正则匹配看是否含有
        flag2 = re.findall('培训',temp_list[5])#从domain属性中用正则匹配看是否含有
        if flag1 or flag2:
            final_result.append(result_list[i])
    print('after_filted_by_domain总数', len(final_result))  # 筛选后一共2609组数据
    return final_result



def information_stat_by_province(result_list):
    """
    按照城市信息统计语料中的信息
    :param result_list:  从excel中读取并处理后的信息
    :return:返回collections.defaultdict格式的一对多字典，即一个键对应多个键值，此键为城市名称，键值为对应的语料索引
    """
    provinces=[]#获取全部的省份
    d = defaultdict(list)#用来存放城市语料统计信息
    for i in range(len(result_list)):  # 依次获取每组数据
        temp_list = result_list[i].split(
            '$')  # 将每一组数据split并重新存为list格式,每行5个属性，分别为0index，1name，2province，3city，4tele，5domain
        provinces.append(temp_list[2])#获取每条语料的省份信息
    for i,each in enumerate(provinces):#对每条语料的省份信息添加index,并统计每个城市对应的语料索引
        d[each].append(i)
    return d


def save_baseon_province(dict,result_list):
    """
    按城市保存：按information_stat_by_province返回的一对多的城市统计信息，将数据按照一个city一个sheet存入excel
    :param dict: 每个城市对应的多条语料的索引
    :param result_list: #待处理的数据，主要要跟传入information_stat_by_province的一致
    :return: None
    """
    book=openpyxl.Workbook()#创建一个工作薄
    summary=[]#存放总结用数据的容器
    for key_city in dict:#遍历字典中的所有城市
        data_excel = []  # 待存入表的数据容器
        sheet = book.create_sheet(key_city)  # 创建一个新的sheet并将新的sheet表名称改为城市名
        sheet['A1']='公司名称'#插入表头
        sheet['B1']='公司电话'
        sheet['C1']='数量总计'
        indexs_of_each_city=dict[key_city]#获取字典中每个city对应的所有索引，
        #print(indexs_of_each_city)#一个城市对应的多个索引，以list格式保存
        sheet['D1']=len(indexs_of_each_city)#每个城市对应的sheet上边标出当页数据总量
        summary.append([key_city,len(indexs_of_each_city)])#将总结数据存入代存容器
        for each_index in indexs_of_each_city:#对每个city下的多个索引依次遍历处理
            temp_list=result_list[each_index].split('$')#split为5个属性0index，1name，2province，3city，4tele，5domain
            data_excel.append([temp_list[1],temp_list[4]])#将公司名称与电话号码放入待存入表的容器
        for each in data_excel:#所有数据提取完后，统一存入表中
            sheet.append(each)

    summary_sheet=book.create_sheet('总结')#创建总结sheet
    summary_sheet['A1']='省份'#添加表头
    summary_sheet['B1']='总数'
    for each in summary:
        summary_sheet.append(each)#将总结信息存入sheet
    book.save(save_file_path)

def figure_genertor(excel_file):
    summary=pd.read_excel(excel_file,sheetname='总结',index_col='省份')#读取excle文件中的“总结”sheet，以“省份”对应的列的内容为数据标签
    #print(summary)#读取excel表格数据
    summary['总数'].plot.pie(subplots=True,figsize=(6, 6),autopct='%.2f')#绘制饼图，以“总数”对应的列的内容为数据绘制饼图，同时设置“总数”为饼图标题
    plt.show()
    summary['总数'].plot.bar()#绘制直方图，以“总数”对应的列的内容为数据绘制直方图
    plt.show()


if __name__=='__main__':
    #得到表格数据
    # raw_datas=read_data(data_file_path)#第一步，对原始语料进行预处理，获取raw_data
    # first_time=filter_by_tele(raw_datas)#根据电话号码对数据初步筛选
    # #second_time=filter_by_domain(first_time)#根据经营范围对数据再次筛选
    # print('数据总数',len(first_time))#最终所得的数据总数
    # city_dic=information_stat_by_province(first_time)#根据筛选后的数据获取一对多的city统计dic
    # save_baseon_province(city_dic,first_time)#按照一个城市一个sheet保存数据

    #绘图(可以修改一下生成的excel文件中的‘总结’sheet中的数据在出图)
    figure_genertor(save_file_path)#根据保存的数据绘制饼图